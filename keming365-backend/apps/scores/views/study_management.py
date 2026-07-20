# -*- coding: utf-8 -*-
"""Study management APIs compatible with the legacy sycj.jsp page."""

from io import BytesIO

import openpyxl
from django.db import connection
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


def _params(request):
    return request.query_params if request.method == 'GET' else request.data


def _is_student(user):
    return int(getattr(user, 'type', 5) or 5) == 2


def _base_where(user, curriculum_id='', experiment_id=''):
    where = []
    values = []
    if _is_student(user):
        where.append('s.user_id=%s')
        values.append(str(user.id))
    if curriculum_id:
        where.append('s.curriculum_id=%s')
        values.append(curriculum_id)
    if experiment_id:
        where.append('s.experiment_id=%s')
        values.append(experiment_id)
    return (' WHERE ' + ' AND '.join(where)) if where else '', values


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def study_courses(request):
    """Course dropdown for 学习管理."""
    where, values = _base_where(request.user)
    sql = f"""
        SELECT DISTINCT s.curriculum_id, COALESCE(c.curriculum_name, '')
        FROM tb_experiment_score s
        LEFT JOIN tb_curriculum c ON s.curriculum_id = c.id
        {where}
        ORDER BY COALESCE(c.curriculum_name, '')
    """
    with connection.cursor() as cur:
        cur.execute(sql, values)
        rows = cur.fetchall()
    return Response([
        {'curriculumId': row[0], 'curriculumStr': row[1] or '未知课程'}
        for row in rows if row[0]
    ])


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def study_experiments(request):
    """Experiment dropdown for 学习管理."""
    params = _params(request)
    curriculum_id = params.get('curriculumId') or params.get('cId') or ''
    where, values = _base_where(request.user, curriculum_id=curriculum_id)
    sql = f"""
        SELECT DISTINCT s.experiment_id, COALESCE(e.title, '')
        FROM tb_experiment_score s
        LEFT JOIN tb_experiment e ON s.experiment_id = e.id
        {where}
        ORDER BY COALESCE(e.title, '')
    """
    with connection.cursor() as cur:
        cur.execute(sql, values)
        rows = cur.fetchall()
    return Response([
        {'experimentId': row[0], 'experimentStr': row[1] or '未知实验'}
        for row in rows if row[0]
    ])


def _fetch_score_rows(user, curriculum_id='', experiment_id='', page=1, page_size=10, paginate=True):
    where, values = _base_where(user, curriculum_id=curriculum_id, experiment_id=experiment_id)
    count_sql = f"SELECT COUNT(*) FROM tb_experiment_score s {where}"
    data_sql = f"""
        SELECT s.id,
               COALESCE(u.school_name, ''),
               COALESCE(u.name, s.un, ''),
               COALESCE(c.curriculum_name, ''),
               COALESCE(e.title, ''),
               s.operation_score,
               s.report_score,
               s.score_sum,
               s.experiment_num,
               COALESCE(ci.class_card, '')
        FROM tb_experiment_score s
        LEFT JOIN tb_user u ON s.user_id = u.id
        LEFT JOIN tb_curriculum c ON s.curriculum_id = c.id
        LEFT JOIN tb_experiment e ON s.experiment_id = e.id
        LEFT JOIN tb_class_info ci ON s.class_Id = ci.id
        {where}
        ORDER BY s.create_time DESC
    """
    with connection.cursor() as cur:
        cur.execute(count_sql, values)
        total = cur.fetchone()[0] or 0
        query_values = list(values)
        if paginate:
            offset = (page - 1) * page_size
            data_sql += ' LIMIT %s OFFSET %s'
            query_values.extend([page_size, offset])
        cur.execute(data_sql, query_values)
        rows = cur.fetchall()

    results = []
    for row in rows:
        results.append({
            'id': row[0],
            'schoolStr': row[1] or '',
            'studentName': row[2] or '',
            'curriculumStr': row[3] or '',
            'experimentStr': row[4] or '',
            'operationScore': '' if row[5] is None else str(row[5]),
            'reportScore': '' if row[6] is None else str(row[6]),
            'scoreSum': '' if row[7] is None else str(row[7]),
            'experimentNum': row[8] or 0,
            'classStr': row[9] or '',
        })
    return total, results


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def study_scores(request):
    """Score table for 学习管理."""
    params = _params(request)
    curriculum_id = params.get('curriculumId') or params.get('cId') or ''
    experiment_id = params.get('experimentId') or params.get('eId') or ''
    page = int(params.get('page') or params.get('startPage') or 1)
    page_size = int(params.get('page_size') or params.get('PageSize') or 10)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 100)

    total, rows = _fetch_score_rows(request.user, curriculum_id, experiment_id, page, page_size)
    pages = (total + page_size - 1) // page_size if total else 0
    request.session['StuCJResultMap'] = {'list': rows}
    request.session.modified = True
    return Response({
        'count': total,
        'results': rows,
        'pageInfo': {
            'pageNum': page,
            'pageSize': page_size,
            'pages': pages,
            'total': total,
            'prePage': max(page - 1, 1),
            'nextPage': min(page + 1, pages or 1),
        },
        'list': {
            'rows': rows,
            'pageInfo': {
                'pageNum': page,
                'pages': pages,
                'total': total,
                'prePage': max(page - 1, 1),
                'nextPage': min(page + 1, pages or 1),
                'navigatepageNums': list(range(1, pages + 1)),
            }
        }
    })


def _excel_response(rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成绩明细'
    headers = ['学校', '姓名', '课程', '资源', '操作成绩']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='center', vertical='center')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = align
    for row_idx, item in enumerate(rows, 2):
        values = [item['schoolStr'], item['studentName'], item['curriculumStr'], item['experimentStr'], item['operationScore']]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = align
    for column in ws.columns:
        letter = column[0].column_letter
        ws.column_dimensions[letter].width = min(max(len(str(c.value or '')) for c in column) + 4, 36)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def study_scores_export(request):
    params = request.query_params
    curriculum_id = params.get('curriculumId') or params.get('cId') or ''
    experiment_id = params.get('experimentId') or params.get('eId') or ''
    _, rows = _fetch_score_rows(request.user, curriculum_id, experiment_id, paginate=False)
    return _excel_response(rows, 'study-scores.xlsx')
